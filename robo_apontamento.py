from selenium import webdriver
from datetime import datetime
import pandas as pd
from time import sleep
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoAlertPresentException
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.common.exceptions import StaleElementReferenceException

# Configurações globais
CHROMEDRIVER_PATH = "chromedriver-win64\chromedriver.exe"
SITE_URL = "https://cenegedrj.gpm.srv.br/index.php"
LOGIN_USUARIO = "FABRICIO.GAMA"
SENHA_USUARIO = "115433397*"
file_path = (r"E:\Fabrício\Trabalho\Ceneged\codigo ceneged\RobosCeneged\dados_apontamento_teste - Copia.xlsx")

class Apontamento:
    def __init__(self, file_path):
        """Inicializa o driver e configurações do navegador."""
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)  # Mantém o navegador aberto
        chrome_options.add_argument("--disable-infobars")  # Remove barra de informações
        chrome_options.add_argument("--disable-notifications")  # Desabilita notificações
        chrome_options.add_argument("--start-maximized")  # Tenta iniciar maximizado

        # Carregar a planilha
        self.df = pd.read_excel(file_path)

        self.service = Service(CHROMEDRIVER_PATH)
        self.driver = webdriver.Chrome(service=self.service)
        self.wait = WebDriverWait(self.driver, 20)
        self.log_file = "robo_apontamento_log.txt"
        self._init_log()

    def _init_log(self):
        """Inicializa o arquivo de log."""
        with open(self.log_file, "w") as log:
            log.write(f"Log iniciado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            log.write("=" * 50 + "\n")

    def log(self, message):
        """Registra uma mensagem no arquivo de log."""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_message = f"[{timestamp}] {message}"
        
        # Evitar mensagens duplicadas
        if not hasattr(self, '_last_log_message') or self._last_log_message != log_message:
            with open(self.log_file, "a") as log:
                log.write(log_message + "\n")
            print(log_message)  # Também imprime no console
            self._last_log_message = log_message

    def login(self):
        """Realiza o login no site."""
        try:
            self.driver.get(SITE_URL)
            self.wait.until(EC.presence_of_element_located((By.ID, "idLogin"))).send_keys(LOGIN_USUARIO)
            self.wait.until(EC.presence_of_element_located((By.ID, "idSenha"))).send_keys(SENHA_USUARIO)
            self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "blogin"))).click()
            self.log("Login realizado com sucesso.")
        except TimeoutException:
            self.log("Erro ao realizar o login. Verifique as credenciais ou a conexão.")
            self.driver.quit()
            exit()

    def botao_serviço(self):
        """ Clicar no botão de serviço """
        try:
            # Clicar no botão de serviço
            self.wait.until(EC.element_to_be_clickable((By.ID, "2000"))).click()
            self.log("Botão de serviço clicado.")
        except TimeoutException:
            self.log("Botão de serviço não encontrado.")

    def _acessar_iframes_lateral(self):
        """Troca para os iframes laterais."""
        try:
            # Acessar iframe lateral
            iframe_lateral = self.wait.until(EC.presence_of_element_located((By.ID, "frame_lateral")))
            self.driver.implicitly_wait(5)
            self.driver.switch_to.frame(iframe_lateral)
            self.log("Mudança para o iframe lateral realizada com sucesso.")
            # Clicar nos itens do menu
            self.driver.find_element(By.ID, "jt80").click()
            self.driver.find_element(By.ID, "jt111").click()
            self.driver.find_element(By.ID, "jt112").click()
            self.log("Itens do menu clicados com sucesso.")
            self.driver.switch_to.default_content()  # Voltar ao contexto principal
        except TimeoutException:
            self.log("Erro ao acessar os iframes. Verifique a estrutura da página.")
            # self.driver.quit()
            # exit()

    def _acessar_iframes_central(self):
        """Troca para os iframes centrais."""
        try:
            # Alternar para o contexto padrão antes de acessar o iframe central
            self.driver.switch_to.default_content()
            iframe_central = self.wait.until(EC.presence_of_element_located((By.ID, "frame_central")))
            self.driver.switch_to.frame(iframe_central)
            self.log("Mudança para o iframe central realizada com sucesso.")
        except Exception as e:
            self.log(f"Mudança para o iframe central erro: {e}.")

    def _acessar_iframes_secundarios(self):
        """Acessa os iframes secundários."""
        try:
            self.log("Tentando acessar o iframe secundário...")
            # Verifica se o iframe já está ativo
            if self.driver.find_element(By.ID, "frm_down").is_displayed():
                self.log("Iframe secundário já está ativo.")

                # Acessar iframe secundário
                iframe_servico = self.wait.until(EC.presence_of_element_located((By.ID, "frm_down")))
                self.driver.switch_to.frame(iframe_servico)
                self.log("Mudança para o iframe lateral realizada com sucesso.")
                return  
         
        except TimeoutException:
            self.log("Erro ao acessar os iframes. Verifique a estrutura da página.")
            self.driver.quit()
            exit()

    def preencher_cabecalho(self, row):
        try:

            # Preenchimento dos campos
            self._preencher_com_sugestao("inputString", row["inputString"], "autoSuggestionsList")
            sleep(0.5)
            self._interagir_dropdown("contrato_chosen", row["contrato_chosen"])
            sleep(0.2)
            self._interagir_dropdown("equipe_chosen", row["equipe_chosen"])
            sleep(0.2)
            self._interagir_dropdown("tip_srv_chosen", row["tip_srv_chosen"])
            sleep(0.2)
            self._interagir_dropdown("obras_chosen", row["obras_chosen"])
            sleep(0.2)
            self._interagir_dropdown("cod_irr_chosen", str(row["cod_irr_chosen"]))
            sleep(0.5)

            # Preenchimento de data
            self._preencher_campo_data_hora("dat_srv", row["dat_srv"].strftime("%d/%m/%Y"))
            sleep(0.2)
            self._preencher_campo_data_hora("hr_inic", row["hr_inic"].strftime("%H:%M"))
            sleep(0.2)
            self._preencher_campo_data_hora("dat_srv2", row["dat_srv2"].strftime("%d/%m/%Y"))
            sleep(0.2)
            self._preencher_campo_data_hora("hr_fim", row["hr_fim"].strftime("%H:%M"))
            self.log(f"Cabeçalho preenchido com sucesso: {row}")
        except Exception as e:
            self.log(f"Erro ao preencher o cabeçalho: {row} , {e}")
        finally:
            # Salvar
            apontamento.salvar()
            # Acessar iframe secundário
            apontamento._acessar_iframes_secundarios()

    def preencher_servico(self, row):
        """Preenche os dados do servico."""
        try:
            texto = str(row["serv_chosen"])
    
            # Preencher o campo de texto e verificar se foi bem-sucedido
            self._interagir_dropdown_serviso("serv_chosen", texto)

        except Exception as e:
            self.log(f"Erro ao preencher serv_chosen: {texto}: {e}")
            return  # Sai da função em caso de exceção

        try:
            # Garantir que 'qtd' seja um número válido e formatá-lo com dois dígitos decimais
            if isinstance(row["qtd"], (int, float)):
                valor_qtd = f"{float(row['qtd']):.2f}"  # Formata o número para sempre ter 2 casas decimais
            else:
                self.log(f"Valor inválido para 'qtd': {row['qtd']}")
                raise ValueError(f"Valor inválido para 'qtd': {row['qtd']}")
            
            # Preencher o campo com o valor formatado
            self._preencher_campo_valor("qtd", valor_qtd)
                
        except Exception as e:
            self.log(f"Erro ao preencher os dados {valor_qtd}: {e}")
            self.cancelar()
            self.fechar()

        try:

            sleep(0.5)

            # Incluir o apontamento somente se o serviço foi preenchido com sucesso
            apontamento.incluir()

        except Exception as e:
            self.log(f"Erro ao incluir o apontamento: {e}")

    def _preencher_com_sugestao(self, campo_id, texto, suggestion_list_id):
        """Preenche um campo de texto e clica na sugestão correspondente usando XPath."""
        try:

            # Localizar o campo de texto e inserir o texto
            campo_texto = self.wait.until(EC.element_to_be_clickable((By.ID, campo_id)))
            campo_texto.clear()
            campo_texto.send_keys(texto)
            self.log(f"Texto '{texto}' inserido no campo '{campo_id}'.")

            # Construir XPath diretamente
            xpath = f"//*[@id='{suggestion_list_id}']"
            sugestoes = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))

            self.log(f"Lista de sugestões carregada: {len(sugestoes)} itens encontrados.")
            
            # Clicar na primeira sugestão
            primeira_sugestao = sugestoes[0]
            texto_primeira_sugestao = primeira_sugestao.text.strip()
            self.log(f"Texto da primeira sugestão: '{texto_primeira_sugestao}'.")
         
            primeira_sugestao.click()
            self.log(f"Sugestão '{texto_primeira_sugestao}' clicada com sucesso via XPath.")

        except TimeoutException:
            self.log(f"Erro: Tempo limite ao preencher ou selecionar a sugestão no campo '{campo_id}'.")
        except Exception as e:
            self.log(f"Erro inesperado ao processar a sugestão: {e}")

    def _preencher_e_confirmar_serviso(self, campo, texto):
        """Insere texto em um campo e aguarda para pressionar Enter."""
        try:
            # Converte o texto para string, caso não seja
            textos = str(texto)
            
            # Aguarda até que o campo esteja clicável e recarrega o elemento
            campo = self.wait.until(EC.element_to_be_clickable(campo))
            campo.click()

            sleep(0.3)

            # Limpa o campo antes de preencher (opcional, dependendo do comportamento desejado)
            campo.clear()

            sleep(0.3)

            # Enviar o texto
            campo.send_keys(textos)
            self.log(f"Texto '{textos}' inserido no campo com sucesso.")
            
            # Pressionar Enter
            campo.send_keys(Keys.ENTER)
            self.log("Tecla Enter pressionada.")

        except TimeoutException:
            self.log(f"Erro ao inserir e confirmar o texto '{textos}': Timeout ao aguardar o campo ou o texto.")
        except ValueError as e:
            self.log(f"Erro ao inserir e confirmar o texto '{textos}': {e}")
        except StaleElementReferenceException:
            self.log(f"Elemento obsoleto ao tentar preencher o texto '{textos}'. Tentando novamente...")
            self._preencher_e_confirmar_serviso(campo, texto)  # Tenta novamente
        except Exception as e:
            self.log(f"Erro inesperado ao inserir e confirmar o texto '{textos}': {str(e)}")
            self.cancelar()

    def _preencher_e_confirmar(self, campo, texto):
        """Insere texto em um campo e aguarda para pressionar Enter."""
        try:
            # Verificar se o campo existe e é válido
            if campo is None:
                raise ValueError("Campo é nulo")
            if not hasattr(campo, 'is_enabled') or not hasattr(campo, 'is_displayed'):
                raise ValueError("Campo não é um elemento válido")
            
            # Converte o texto para string, caso não seja
            texto = str(texto)
            
            # Aguarda até que o campo esteja clicável
            campo = self.wait.until(EC.element_to_be_clickable(campo))
            campo.click()

                                            
            # Verificar se o elemento ainda é válido
            if not campo.is_enabled() or not campo.is_displayed():
                raise Exception("Elemento não está mais disponível ou visível.")

            # Limpar o campo
            campo.clear()

            # Enviar o texto
            campo.send_keys(texto)
            self.log(f"Texto '{texto}' inserido no campo com sucesso.")
            
            # Aguardar até que o texto seja refletido no campo
            self.wait.until(lambda driver: campo.get_attribute("value") is not None and texto.lower() in campo.get_attribute("value").lower())
            self.log(f"Confirmação de que o campo contém o texto '{texto}'.")
            
            # Pressionar Enter
            campo.send_keys(Keys.ENTER)
            self.log("Tecla Enter pressionada.")
            
        except TimeoutException:
            self.log(f"Erro ao inserir e confirmar o texto '{texto}': Timeout ao aguardar o campo ou o texto.")
        except ValueError as e:
            self.log(f"Erro ao inserir e confirmar o texto '{texto}': {e}")
        except Exception as e:
            self.log(f"Erro inesperado ao inserir e confirmar o texto '{texto}': {str(e)}")
            self.cancelar()

    def _interagir_dropdown_serviso(self, dropdown_id, texto):
        """Interage com um dropdown customizado."""
        #max_tentativas = 3  # Número máximo de tentativas
        #for tentativa in range(max_tentativas):
        try:
            #self.log(f"Tentativa {tentativa + 1} de interagir com o dropdown '{dropdown_id}'.")

            # Aguardar até que o dropdown esteja clicável

            sleep(0.5)
            dropdown_service = self.wait.until(EC.element_to_be_clickable((By.ID, dropdown_id)))
            dropdown_service.click()
            
            self.log(f"Dropdown '{dropdown_id}' aberto com sucesso.")

            # Localiza o campo de busca dentro do dropdown
            search_box = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, f"#{dropdown_id} div.chosen-search input"))
            )
            sleep(0.5)            
            # Preenche o texto e confirma com Enter
            self._preencher_e_confirmar_serviso(search_box, texto)

            # Verifica se o texto foi inserido corretamente
            if search_box.get_attribute("value") == texto:
                self.log(f"Texto '{texto}' preenchido com sucesso no dropdown '{dropdown_id}'.")
                return True  # Retorna True se o texto foi preenchido corretamente
            else:
                self.log(f"Falha ao preencher o texto '{texto}' no dropdown '{dropdown_id}'.")
                return False  # Retorna False se o texto não foi preenchido corretamente

        except StaleElementReferenceException:
            self.log(f"Elemento obsoleto na tentativa tentativa + 1. Tentando novamente...")
            #continue  # Tenta novamente em caso de elemento obsoleto
        except TimeoutException:
            self.log(f"Timeout ao interagir com o dropdown '{dropdown_id}' na tentativa tentativa + 1.")
            #continue  # Tenta novamente em caso de timeout
        except Exception as e:
            self.log(f"Erro inesperado ao interagir com o dropdown '{dropdown_id}': {e}")
            return False  # Retorna False em caso de exceção

        # self.log(f"Falha após {max_tentativas} tentativas de interagir com o dropdown '{dropdown_id}'.")
        # return False  # Retorna False se todas as tentativas falharem
    def _interagir_dropdown(self, dropdown_id, texto):
        """Interage com um dropdown customizado."""
        try:

            # Abre o dropdown
            self.wait.until(EC.element_to_be_clickable((By.ID, dropdown_id ))).click()
            
            
            self.log(f"Dropdown '{dropdown_id}' aberto com sucesso.")

            # Localiza o campo de busca dentro do dropdown
            search_box = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, f"#{dropdown_id} div.chosen-search input"))
            )

            sleep(0.5)
            
            # Preenche o texto e confirma com Enter
            self._preencher_e_confirmar(search_box, texto)
        except TimeoutException:
            self.log(f"Erro ao interagir com o dropdown '{dropdown_id}'.")

    def _preencher_campo_valor(self, campo_id, valor):
        """Preenche um campo de texto ou data/hora."""
        try:
            campo = self.wait.until(EC.element_to_be_clickable((By.ID, campo_id)))
            self.driver.execute_script("arguments[0].value = arguments[1];", campo, valor)
            self.log(f"Campo '{campo_id}' preenchido com valor '{valor}'.")
        except TimeoutException:
            self.log(f"Erro ao preencher o campo '{campo_id}'.")

    def _preencher_campo_data_hora(self, campo_id, valor):
        """Preenche um campo de texto ou data/hora."""
        try:
            campo = self.wait.until(EC.element_to_be_clickable((By.ID, campo_id)))
            self.driver.execute_script("arguments[0].value = arguments[1];", campo, valor)
            self.log(f"Campo '{campo_id}' preenchido com valor '{valor}'.")
        except TimeoutException:
            self.log(f"Erro ao preencher o campo '{campo_id}'.")

    def identificar_servicos_duplicados(self):
        """
        Identifica serviços duplicados para a mesma combinação de obra, data e equipe.
        Se encontrar duplicatas, chama a função fechar().
        
        Returns:
            bool: True se encontrou duplicatas, False caso contrário
        """
        try:
            # Converte todos os valores para string antes de concatenar
            self.df['chave'] = (
                self.df['obras_chosen'].astype(str) + '_' + 
                self.df['dat_srv'].astype(str) + '_' + 
                self.df['equipe_chosen'].astype(str)
            )
            # Verifica duplicatas considerando a chave e o serviço
            self.df['duplicado_servico'] = self.df.duplicated(
                subset=['chave', 'serv_chosen'], 
                keep=False
            )
            # Remove a coluna chave temporária
            self.df.drop('chave', axis=1, inplace=True)
            
            # Verifica se há algum serviço duplicado
            if self.df['duplicado_servico'].any():
                self.log("Serviços duplicados encontrados.")
                self.exibir_servicos_repetidos()  # Chama a função exibir_servicos_repetidos
                self.fechar()  # Chama a função fechar() se encontrar duplicatas
                return True
            else:
                self.log("Nenhum serviço duplicado encontrado.")
                return False
                
        except Exception as e:
            self.log(f"Erro ao identificar serviços duplicados: {str(e)}")
            return False
        
    def exibir_servicos_repetidos(self):
        """Exibe os serviços repetidos de forma formatada e organizada"""
        try:
            # Obtém os dados repetidos
            df_repetidos = self._get_dados_repetidos()
            
            if df_repetidos.empty:
                print("\nNenhum serviço duplicado encontrado.")
                return
            
            print("\n=== SERVIÇOS DUPLICADOS ENCONTRADOS ===")
            
            # Exibe os serviços repetidos aonde _ é quebra de linha e row é uma linha
            for _, row in df_repetidos.iterrows():
                print(f"\nObra: {row['obra']}")
                print(f"Data: {row['data']}")
                print(f"Equipe: {row['equipe']}")
                print("Serviços repetidos:")
                
                for servico, quantidade in row['vezes'].items():
                    print(f"  - Código: {servico} (Repetido {quantidade} vezes)")
                    self.log(f"Serviços repetidos: Obra: {row['obra']}\n - Data: {row['data']}\n - Equipe: {row['equipe']}\n  - Código: {servico} (Repetido {quantidade} vezes)")
                    
            print("\n" + "="*40 + "\n")
            return True
            
        except Exception as e:
            self.log(f"Erro ao exibir serviços duplicados: {str(e)}")
            return False

    def _get_dados_repetidos(self):
        """Método auxiliar para obter os dados de serviços repetidos"""
        grupos = self.df.groupby(['obras_chosen', 'dat_srv', 'equipe_chosen'])
        
        resultados = []
        for nome, grupo in grupos:
            contagem = grupo['serv_chosen'].value_counts()
            repetidos = contagem[contagem > 1]
            
            if not repetidos.empty:
                resultados.append({
                    'obra': nome[0],
                    'data': nome[1],
                    'equipe': nome[2],
                    'vezes': repetidos.to_dict()
                })
        
        return pd.DataFrame(resultados)

    def fechar(self):
        """Finaliza o WebDriver."""
        try:
            self.driver.quit()
            self.log("WebDriver finalizado com sucesso.")
        except Exception as e:
            self.log(f"Erro ao finalizar o WebDriver: {e}")

    def salvar(self):
        """Salva o formulário."""
        try:
            salvar = self.wait.until(EC.element_to_be_clickable((By.ID, "idSubmit")))
            salvar.click()
            self.log("Formulário salvo com sucesso.")
            # Pausa breve
            sleep(0.1)
        except TimeoutException:
            self.log("Erro ao salvar o formulário.")

    def incluir(self):
        """Clica no botão Incluir."""
        try:

            # Aguarda até que o botão esteja clicável
            botao_incluir = self.wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input.dt-button.btn_green"))
            )

            sleep(0.2)

            botao_incluir.click()
            self.log("Botão 'Incluir' clicado com sucesso.")
        except TimeoutException:
            self.log("Erro ao clicar no botão 'Incluir'. Verifique se ele está visível na página.")
            self.fechar()

    def finalizar(self):
        """Finaliza o ciclo clicando no botão 'finalizar'."""
        try:
            # Garantir que estamos no iframe central
            self._acessar_iframes_central()

            # Localizar e clicar no botão "Finalizar"
            botao_finalizar = self.wait.until(EC.element_to_be_clickable((By.ID, "idSubmit")))
            botao_finalizar.click()
            self.log("Botão 'Finalizar' clicado com sucesso.")

            self._alerta_auto()
            
        except Exception as e:
            self.log(f"Erro ao finalizar: {e}")
    def cancelar(self):
        """Cancela o ciclo clicando no botão 'Cancelar'."""
        try:
            # Garantir que estamos no iframe central
            self._acessar_iframes_central()

            # Localizar e clicar no botão "Cancelar"
            botao_cancelar = self.wait.until(EC.element_to_be_clickable((By.ID, "idCancel")))
            botao_cancelar.click()
            self.log("Botão 'Cancelar' clicado com sucesso.")

            self._alerta_auto()
            self.fechar()
            
        except Exception as e:
            self.log(f"Erro ao cancelar: {e}")

    def _alerta_auto(self):
        """Fecha os alertas automaticamente."""
        try:
            # Aceitar o alerta automaticamente
            self.wait.until(EC.alert_is_present())  # Aguarda a presença do alerta
            self.driver.switch_to.alert.accept()
            self.log("Alerta e concluido aceito automaticamente.")
        except NoAlertPresentException:
            self.log("Nenhum alerta encontrado ao clicar em finalizar.") 

    def salvar_planilha_com_formatacao(self, df, file_path):
        """
        Salva a planilha mantendo a formatação de data/hora.
        """
        try:
            # Carregar a planilha existente
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Limpar os dados antigos
            sheet.delete_rows(2, sheet.max_row)

            # Adicionar os dados do DataFrame (excluindo o cabeçalho)
            for row in dataframe_to_rows(df, index=False, header=False):  # header=False evita duplicar o cabeçalho
                sheet.append(row)

            # Ajustar formatação para colunas de data/hora (exemplo: colunas 7, 8, 9, 10)
            for col in [7, 8, 9, 10]:  # Substitua pelas colunas de data/hora no índice correto
                for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        c.number_format = 'DD/MM/YYYY' if col in [7, 9] else 'HH:MM'

            # Salvar novamente
            workbook.save(file_path)
            self.log("Planilha salva mantendo a formatação de data/hora.")

        except Exception as e:
            self.log(f"Erro ao salvar planilha mantendo a formatação: {e}")   

    def executar_planilha(self):
        """Executa as entradas da planilha com base em cabeçalhos e serviços."""
        try:


            # Verificar se as colunas 'status' e 'projeto' existem
            if 'status' not in self.df.columns or 'obras_chosen' not in self.df.columns:
                self.log("Erro: A planilha não possui as colunas 'status' ou 'projeto'.")
                return

            # Garantir que a coluna 'status' existe e preencher com None, caso esteja vazia
            self.df['status'] = self.df['status'].fillna('')

            # Definir as colunas de cabeçalho
            colunas_cabecalho = [
                "obras_chosen", "contrato_chosen", "equipe_chosen",
                "tip_srv_chosen", "inputString", "cod_irr_chosen",
                "dat_srv", "hr_inic", "dat_srv2", "hr_fim"
            ]

            # Inicializar o cabeçalho anterior como vazio
            ultimo_cabecalho = []

            # Criar uma cópia filtrada para processar apenas linhas pendentes
            df_pendentes = self.df[self.df['status'] != 'ok'].copy()

            # Iterar pelas linhas da planilha pendente
            for index, row in df_pendentes.iterrows():
                # Verificar se a coluna 'obras_chosen' está vazia
                if pd.isnull(row['obras_chosen']) or str(row['obras_chosen']).strip() == '':
                    self.log(f"Linha {index} ignorada: coluna 'projeto' está vazia.")
                    continue

                self.log(f"Processando linha {index}...")

                # Extrair o cabeçalho atual
                cabecalho_atual = {col: row[col] for col in colunas_cabecalho}

                # Verificar se o cabeçalho mudou em relação ao último
                if cabecalho_atual != ultimo_cabecalho:
                    # Finalizar o ciclo anterior, se necessário
                    if ultimo_cabecalho:
                        self.log(f"Ciclo: {index} para o cabeçalho anterior {index - 1}.")
                        self.finalizar()
                        self.log(f"Ciclo finalizado para o cabeçalho anterior na linha {index}.")

                    # Atualizar o cabeçalho
                    try:
                        self.preencher_cabecalho(row)
                        ultimo_cabecalho = cabecalho_atual
                        self.log(f"Novo cabeçalho processado na linha {index}.")
                    except Exception as e:
                        self.log(f"Erro ao preencher o cabeçalho na linha {index}: {e}")
                        continue

                # Preencher o serviço correspondente à linha atual
                try:
                    self.preencher_servico(row)
                    self.log(f"Linha {index} processada com sucesso.")

                    # Atualizar o status na versão original do DataFrame
                    self.df.at[row.name, 'status'] = 'ok'
                    self.log(f"Status ok atualizado na linha {index}.")

                    # Salvar a planilha atualizada
                    self.salvar_planilha_com_formatacao(self.df, file_path)
                    self.log("Planilha salva com status atualizado.")
                    sleep(0.5)
                except Exception as e:
                    self.log(f"Erro ao preencher o serviço na linha {index}: {e}")
                    continue

            # Finalizar o último ciclo após o loop
            self.finalizar()
            self.log("Último ciclo finalizado com sucesso.")

            # Salvar a planilha completa ao final
            self.salvar_planilha_com_formatacao(self.df, file_path)
            self.log("Planilha salva com todos os status atualizados.")

        except Exception as e:
            self.log(f"Erro ao processar a planilha: {e}")

# Execução do Script
if __name__ == "__main__":
    apontamento = Apontamento(file_path)
    try:
        apontamento.identificar_servicos_duplicados()
        apontamento.login() 
        apontamento.botao_serviço()
        apontamento._acessar_iframes_lateral()
        apontamento._acessar_iframes_central()
        apontamento.executar_planilha()
    except Exception as e:
        apontamento.log(f"Erro inesperado: {e}")
    finally:
        apontamento.fechar()
        apontamento.log("Script finalizado.")