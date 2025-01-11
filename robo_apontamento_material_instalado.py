from selenium import webdriver
from datetime import datetime
import pandas as pd
from time import sleep
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoAlertPresentException
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurações globais
CHROMEDRIVER_PATH = "chromedriver-win64\chromedriver.exe"
SITE_URL = "https://cenegedrj.gpm.srv.br/index.php"
LOGIN_USUARIO = "fabricio.gama"
SENHA_USUARIO = "11543339735*"

class Apontamento:
    def __init__(self):
        """Inicializa o driver e configurações do navegador."""
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)  # Mantém o navegador aberto
        chrome_options.add_argument("--disable-infobars")  # Remove barra de informações
        chrome_options.add_argument("--disable-notifications")  # Desabilita notificações
        chrome_options.add_argument("--start-maximized")  # Tenta iniciar maximizado

        self.service = Service(CHROMEDRIVER_PATH)
        self.driver = webdriver.Chrome(service=self.service)
        self.wait = WebDriverWait(self.driver, 20)
        self.log_file = "robo_apontamento_log.txt"
        self._init_log()

    def _init_log(self):
        """Inicializa o arquivo de log."""
        with open(self.log_file, "w") as log:
            log.write(f"Log iniciado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            log.write("=" * 50 + "\n")

    def log(self, message):
        """Registra uma mensagem no arquivo de log."""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_message = f"[{timestamp}] {message}"
        
        # Evitar mensagens duplicadas
        if not hasattr(self, '_last_log_message') or self._last_log_message != log_message:
            with open(self.log_file, "a") as log:
                log.write(log_message + "\n")
            print(log_message)  # Também imprime no console
            self._last_log_message = log_message

    def login(self):
        """Realiza o login no site."""
        try:
            self.driver.get(SITE_URL)
            self.wait.until(EC.presence_of_element_located((By.ID, "idLogin"))).send_keys(LOGIN_USUARIO)
            self.wait.until(EC.presence_of_element_located((By.ID, "idSenha"))).send_keys(SENHA_USUARIO)
            self.wait.until(EC.element_to_be_clickable((By.CLASS_NAME, "blogin"))).click()
            self.log("Login realizado com sucesso.")
        except TimeoutException:
            self.log("Erro ao realizar o login. Verifique as credenciais ou a conexão.")
            self.driver.quit()
            exit()

    def botao_serviço(self):
        """ Clicar no botão de serviço """
        try:
            # Clicar no botão de serviço
            self.wait.until(EC.element_to_be_clickable((By.ID, "2000"))).click()
            self.log("Botão de serviço clicado.")
        except TimeoutException:
            self.log("Botão de serviço não encontrado.")

    def botao_material_instalado(self):
        """ Clicar no botão de material instalado """
        try:
            # Clicar no botão de materia
            self.wait.until(EC.element_to_be_clickable ((By.ID, "aba3"))).click()
            self.log("Botão de material instalado clicado.")
        except TimeoutException:
            self.log("Botão de material instalado não encontrado.")

    def _acessar_iframes_lateral(self):
        """Troca para os iframes laterais."""
        try:
            # Acessar iframe lateral
            iframe_lateral = self.wait.until(EC.presence_of_element_located((By.ID, "frame_lateral")))
            self.driver.implicitly_wait(5)
            self.driver.switch_to.frame(iframe_lateral)
            self.log("Mudança para o iframe lateral realizada com sucesso.")
            # Clicar nos itens do menu
            self.driver.find_element(By.ID, "jt80").click()
            self.driver.find_element(By.ID, "jt111").click()
            self.driver.find_element(By.ID, "jt112").click()
            self.log("Itens do menu clicados com sucesso.")
            self.driver.switch_to.default_content()  # Voltar ao contexto principal
        except TimeoutException:
            self.log("Erro ao acessar os iframes. Verifique a estrutura da página.")
            # self.driver.quit()
            # exit()

    def _acessar_iframes_central(self):
        """Troca para os iframes centrais."""
        try:
            # Alternar para o contexto padrão antes de acessar o iframe central
            self.driver.switch_to.default_content()
            iframe_central = self.wait.until(EC.presence_of_element_located((By.ID, "frame_central")))
            self.driver.switch_to.frame(iframe_central)
            self.log("Mudança para o iframe central realizada com sucesso.")
        except Exception as e:
            self.log(f"Mudança para o iframe central erro: {e}.")

    def _acessar_iframes_secundarios(self):
        """Acessa os iframes secundários."""
        try:
            self.log("Tentando acessar o iframe secundário...")

            # Esperar que o iframe 'frm_down' seja adicionado ao DOM
            self.wait.until(
                EC.frame_to_be_available_and_switch_to_it((By.ID, "frm_down"))
            )

            self.log("Mudança para o iframe 'frm_down' realizada com sucesso.")

        except TimeoutException:
            self.log("Erro: Iframe 'frm_down' não foi encontrado no tempo esperado.")
        except Exception as e:
            self.log(f"Erro inesperado ao acessar os iframes: {e}")

    def preencher_cabecalho(self, row):
        try:
                  
            self._preencher_com_sugestao("inputString", row["inputString"], "autoSuggestionsList")
            # Pausa breve
            sleep(0.5)
            self._interagir_dropdown("contrato_chosen", row["contrato_chosen"])
            # Pausa breve
            sleep(0.1)
            self._interagir_dropdown("equipe_chosen", row["equipe_chosen"])
            # Pausa breve
            sleep(0.1)
            self._interagir_dropdown("tip_srv_chosen", row["tip_srv_chosen"])
            # Pausa breve
            sleep(0.1)
            self._interagir_dropdown("obras_chosen", row["obras_chosen"])
            # Pausa breve
            sleep(0.1)
            self._interagir_dropdown("cod_irr_chosen", str(row["cod_irr_chosen"]))
            # Pausa breve
            sleep(0.1)

            # Preenchimento de data
            self._preencher_campo_data_hora("dat_srv", row["dat_srv"].strftime("%d/%m/%Y"))
            # Pausa breve
            sleep(0.1)
            self._preencher_campo_data_hora("hr_inic", row["hr_inic"].strftime("%H:%M"))
            # Pausa breve
            sleep(0.1)
            self._preencher_campo_data_hora("dat_srv2", row["dat_srv2"].strftime("%d/%m/%Y"))
            # Pausa breve
            sleep(0.1)
            self._preencher_campo_data_hora("hr_fim", row["hr_fim"].strftime("%H:%M"))
            self.log("Cabeçalho preenchido com sucesso.")
        except Exception as e:
            self.log(f"Erro ao preencher o cabeçalho: {e}")
        finally:
            # salvar
            apontamento.salvar()
            # Botão material instalado
            apontamento.botao_material_instalado()
            # Acessar iframe secundário
            apontamento._acessar_iframes_secundarios()
            # Associar equipe
            apontamento._associar_equipe()
            # Acessar iframe secundário
            apontamento._acessar_iframes_secundarios()

    def preencher_material_instalado(self, row):
        try:
            # self._preencher_e_confirmar("id_listGrupo_chosen", str(row["listGrupo_chosen"]))
            # # Pausa breve                
            sleep(0.2)
            self._interagir_dropdown("listMater_chosen", str(row["listMater_chosen"]))
            # Pausa breve                
            sleep(0.2)

                        # Garantir que 'qtd' seja um número válido e formatá-lo com dois dígitos decimais
            if isinstance(row["mater"], (int, float)):
                valor_qtd = f"{float(row['mater']):.2f}"  # Formata o número para sempre ter 2 casas decimais
            else:
                raise ValueError(f"Valor inválido para 'mater': {row['mater']}")

            # Preencher o campo com o valor formatado
            self._preencher_campo_data_hora("id_mater", valor_qtd)
            # Pausa breve                
            sleep(0.2)

            self.incluir()

        except TimeoutException:
            self.log(f"Erro: Tempo limite ao preencher ou selecionar a sugestão no campo '{row}'.")
        except Exception as e:
            self.log(f"Erro inesperado ao processar a sugestão: {e}")

    def _preencher_com_sugestao(self, campo_id, texto, suggestion_list_id):
        """Preenche um campo de texto e clica na sugestão correspondente usando XPath."""
        try:
            # Localizar o campo de texto e inserir o texto
            campo_texto = self.wait.until(EC.element_to_be_clickable((By.ID, campo_id)))
            campo_texto.clear()
            campo_texto.send_keys(texto)
            self.log(f"Texto '{texto}' inserido no campo '{campo_id}'.")

            # Construir XPath diretamente
            xpath = f"//*[@id='{suggestion_list_id}']"
            sugestoes = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))

            self.log(f"Lista de sugestões carregada: {len(sugestoes)} itens encontrados.")
            sleep(0.1)
            # Clicar na primeira sugestão
            primeira_sugestao = sugestoes[0]
            texto_primeira_sugestao = primeira_sugestao.text.strip()
            self.log(f"Texto da primeira sugestão: '{texto_primeira_sugestao}'.")
            sleep(0.1)
            primeira_sugestao.click()
            self.log(f"Sugestão '{texto_primeira_sugestao}' clicada com sucesso via XPath.")

        except TimeoutException:
            self.log(f"Erro: Tempo limite ao preencher ou selecionar a sugestão no campo '{campo_id}'.")
        except Exception as e:
            self.log(f"Erro inesperado ao processar a sugestão: {e}")

    def _preencher_e_confirmar(self, campo, texto):
        """Insere texto em um campo e aguarda para pressionar Enter."""
        try:

            # Aguarda até que o campo esteja clicável
            campo = self.wait.until(EC.element_to_be_clickable(campo))
            
            # Enviar o texto
            campo.clear()  # Limpa o campo antes de inserir o texto
            campo.send_keys(texto)
            self.log(f"Texto '{texto}' inserido no campo com sucesso.")

            # Aguarda um pequeno intervalo para que o dropdown processe o texto
            self.wait.until(lambda driver: texto.lower() in campo.get_attribute("value").lower())
            self.log(f"Confirmação de que o campo contém o texto '{texto}'.")

            # Pressionar Enter
            campo.send_keys(Keys.ENTER)

            # Pausa breve
            sleep(0.3)
            
            self.log("Tecla Enter pressionada.")
        except TimeoutException:
            self.log(f"Erro ao inserir e confirmar o texto '{texto}'.")

    def _interagir_dropdown(self, dropdown_id, texto):
        """Interage com um dropdown customizado."""
        try:

            # Abre o dropdown
            dropdown = self.wait.until(EC.element_to_be_clickable((By.ID, dropdown_id)))
            dropdown.click()
            self.log(f"Dropdown '{dropdown_id}' aberto com sucesso.")

            # Localiza o campo de busca dentro do dropdown
            search_box = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, f"#{dropdown_id} div.chosen-search input"))
            )

            sleep(0.5)
            
            # Preenche o texto e confirma com Enter
            self._preencher_e_confirmar(search_box, texto)
        except TimeoutException:
            self.log(f"Erro ao interagir com o dropdown '{dropdown_id}'.")

    def _preencher_campo_data_hora(self, campo_id, valor):
        """Preenche um campo de texto ou data/hora."""
        try:
            campo = self.wait.until(EC.element_to_be_clickable((By.ID, campo_id)))
            self.driver.execute_script("arguments[0].value = arguments[1];", campo, valor)
            self.log(f"Campo '{campo_id}' preenchido com valor '{valor}'.")
        except TimeoutException:
            self.log(f"Erro ao preencher o campo '{campo_id}'.")

    def _associar_equipe(self):
        """Associar a equipe ao apontamento."""
        try:

            xpath = "/html/body/div[1]/p[3]/button"
            # Aguarda até que o botão esteja clicável
            botao_associar = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, xpath))
            )
            sleep(0.2)

            botao_associar.click()
            self.log("Botão 'Associar equipe' clicado com sucesso.")

            # alerta_auto
            self._alerta_auto()
        except TimeoutException:    
            self.log("Erro ao clicar no botão 'Associar equipe'.")

    def fechar(self):
        """Finaliza o WebDriver."""
        try:
            self.driver.quit()
            self.log("WebDriver finalizado com sucesso.")
        except Exception as e:
            self.log(f"Erro ao finalizar o WebDriver: {e}")

    def salvar(self):
        """Salva o formulário."""
        try:
            salvar = self.wait.until(EC.element_to_be_clickable((By.ID, "idSubmit")))
            salvar.click()
            self.log("Formulário salvo com sucesso.")
        except TimeoutException:
            self.log("Erro ao salvar o formulário.")

    def incluir(self):
        """Clica no botão Incluir."""
        try:
            # Aguarda até que o botão esteja clicável
            botao_incluir = self.wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "input.dt-button.btn_green"))
            )
            sleep(0.2)

            botao_incluir.click()
            self.log("Botão 'Incluir' clicado com sucesso.")
        except TimeoutException:
            self.log("Erro ao clicar no botão 'Incluir'. Verifique se ele está visível na página.")

    def finalizar(self):
        """Finaliza o ciclo clicando no botão 'finalizar'."""
        try:
            # Garantir que estamos no iframe central
            self._acessar_iframes_central()

            # Localizar e clicar no botão "Finalizar"
            botao_finalizar = self.wait.until(EC.element_to_be_clickable((By.ID, "idSubmit")))
            botao_finalizar.click()
            self.log("Botão 'Finalizar' clicado com sucesso.")

            self._alerta_auto()
            
        except Exception as e:
            self.log(f"Erro ao finalizar: {e}")

    def _alerta_auto(self):
        """Fecha os alertas automaticamente."""
        try:
            # Aceitar o alerta automaticamente
            self.wait.until(EC.alert_is_present())  # Aguarda a presença do alerta
            self.driver.switch_to.alert.accept()
            self.log("Alerta e concluido aceito automaticamente.")
        except NoAlertPresentException:
            self.log("Nenhum alerta encontrado ao clicar em finalizar.") 

    def salvar_planilha_com_formatacao(self, df, file_path):
        """
        Salva a planilha mantendo a formatação de data/hora.
        """
        try:
            # Carregar a planilha existente
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Limpar os dados antigos
            sheet.delete_rows(2, sheet.max_row)

            # Adicionar os dados do DataFrame (excluindo o cabeçalho)
            for row in dataframe_to_rows(df, index=False, header=False):  # header=False evita duplicar o cabeçalho
                sheet.append(row)

            # Ajustar formatação para colunas de data/hora (exemplo: colunas 7, 8, 9, 10)
            for col in [7, 8, 9, 10]:  # Substitua pelas colunas de data/hora no índice correto
                for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        c.number_format = 'DD/MM/YYYY' if col in [7, 9] else 'HH:MM'

            # Salvar novamente
            workbook.save(file_path)
            self.log("Planilha salva mantendo a formatação de data/hora.")

        except Exception as e:
            self.log(f"Erro ao salvar planilha mantendo a formatação: {e}")   

    def executar_planilha(self, file_path):
        """Executa as entradas da planilha com base em cabeçalhos e serviços."""
        try:
            # Carregar a planilha
            df = pd.read_excel(file_path)

            # Verificar se a coluna 'status' existe
            if 'status' not in df.columns:
                self.log("Erro: A planilha não possui a coluna 'status'.")
                return

            # Garantir que a coluna 'status' existe e preencher com None, caso esteja vazia
            df['status'] = df['status'].fillna('')

            # Definir as colunas de cabeçalho
            colunas_cabecalho = [
                "obras_chosen", "contrato_chosen", "equipe_chosen",
                "tip_srv_chosen", "inputString", "cod_irr_chosen",
                "dat_srv", "hr_inic", "dat_srv2", "hr_fim"
            ]

            # Inicializar o cabeçalho anterior como vazio
            ultimo_cabecalho = []

            # Iterar pelas linhas da planilha
            try:
                for index, row in df.iterrows():
                    # Verificar se o status é 'ok'; se sim, pula a linha
                    if row['status'] == 'ok':
                        continue

                    self.log(f"Processando linha {index}...")

                    # Extrair o cabeçalho atual
                    cabecalho_atual = {col: row[col] for col in colunas_cabecalho}

                    # Verificar se o cabeçalho mudou em relação ao último
                    if cabecalho_atual != ultimo_cabecalho:
                        # Se não for o primeiro ciclo, finalizar o ciclo anterior
                        if ultimo_cabecalho:
                            self.log(f"Ciclo: {index} para o cabeçalho anterior {index - 1}.")
                            self.finalizar()
                            self.log(f"Ciclo finalizado em {index} para o cabeçalho anterior {index - 1}.")

                        try:
                            self.preencher_cabecalho(row)
                            ultimo_cabecalho = cabecalho_atual
                            self.log(f"Novo cabeçalho processado {index}: {cabecalho_atual}")
                        except Exception as e:
                            self.log(f"Erro ao preencher o cabeçalho na linha {index}: {e}")
                            continue  # Pula para a próxima linha em caso de erro

                    # Preencher o material retirado correspondente à linha atual
                    try:
                        self.preencher_material_instalado(row)
                        self.log(f"Linha {index} processada com sucesso.")

                        # Atualizar o status para 'ok'
                        df.at[index, 'status'] = 'ok'

                        # Salvar a planilha atualizada
                        self.salvar_planilha_com_formatacao(df, file_path)
                        self.log("Planilha atualizada com status 'ok'.")
                    except Exception as e:
                        self.log(f"Erro ao preencher o serviço na linha {index}: {e}")
                        continue

            except Exception as e:
                self.log(f"Erro ao iterar pela planilha: {e}")

            # Finalizar o último ciclo após o loop
            self.finalizar()
            self.log("Último ciclo finalizado com sucesso.")

            # Salvar a planilha atualizada
            self.salvar_planilha_com_formatacao(df, file_path)
            self.log("Planilha atualizada com status 'ok'.")

        except Exception as e:
            self.log(f"Erro ao processar a planilha: {e}")

# Execução do Script
if __name__ == "__main__":
    apontamento = Apontamento()
    try:
        apontamento.login() 
        apontamento.botao_serviço()
        apontamento._acessar_iframes_lateral()
        apontamento._acessar_iframes_central()
        apontamento.executar_planilha("dados_apontamento_material_instalado_teste.xlsx")
    except Exception as e:
        apontamento.log(f"Erro inesperado: {e}")
    finally:
        apontamento.fechar()
        apontamento.log("Script finalizado.")